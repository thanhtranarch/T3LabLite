# -*- coding: utf-8 -*-
"""
Sheet Manager - Excel Import/Export Service

Copyright © Dang Quoc Truong (DQT)
"""

import os
import sys


class ExcelService(object):
    """Handle Excel import/export operations"""

    # Columns written by the dict-based export_sheets (used by ManaSheets).
    # The first six match export_sheets_to_excel exactly so a file produced
    # here is still readable by the positional import_sheets_from_excel; the
    # trailing "ID" column is a stable match key for round-trip import, so the
    # user can freely edit sheet number/name and still have rows map back.
    _EXPORT_COLUMNS = [
        ("Sheet Number", "sheet_number"),
        ("Sheet Name",   "sheet_name"),
        ("Designed By",  "designed_by"),
        ("Checked By",   "checked_by"),
        ("Drawn By",     "drawn_by"),
        ("Approved By",  "approved_by"),
        ("ID",           "id"),
    ]

    # Header text (stripped/lowercased) -> dialog field key, for import_sheets.
    _HEADER_TO_KEY = {
        "sheet number": "sheet_number",
        "sheet name":   "sheet_name",
        "designed by":  "designed_by",
        "checked by":   "checked_by",
        "drawn by":     "drawn_by",
        "approved by":  "approved_by",
        "id":           "id",
    }

    def __init__(self):
        self.excel_available = self._check_excel()
    
    def _check_excel(self):
        """Check if Excel libraries are available"""
        try:
            import clr
            clr.AddReference('Microsoft.Office.Interop.Excel')
            return True
        except:
            # Try openpyxl as fallback
            try:
                import openpyxl
                return True
            except:
                return False
    
    def export_sheets_to_excel(self, sheet_models, filepath):
        """Export sheet list to Excel"""
        try:
            # Try openpyxl first (better for PyRevit)
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet List"
            
            # Header style
            header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            
            # Headers
            headers = [
                "Sheet Number",
                "Sheet Name", 
                "Designed By",
                "Checked By",
                "Drawn By",
                "Approved By"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Data rows
            for row, sheet_model in enumerate(sheet_models, 2):
                ws.cell(row=row, column=1, value=sheet_model.sheet_number)
                ws.cell(row=row, column=2, value=sheet_model.sheet_name)
                ws.cell(row=row, column=3, value=sheet_model.designed_by)
                ws.cell(row=row, column=4, value=sheet_model.checked_by)
                ws.cell(row=row, column=5, value=sheet_model.drawn_by)
                ws.cell(row=row, column=6, value=sheet_model.approved_by)
            
            # Auto-size columns
            for col in range(1, 7):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
            
            # Save
            wb.save(filepath)
            return True
            
        except ImportError:
            # Fallback: CSV export
            return self._export_to_csv(sheet_models, filepath.replace('.xlsx', '.csv'))
        except Exception as e:
            print("Error exporting to Excel: {}".format(str(e)))
            import traceback
            traceback.print_exc()
            return False
    
    def _export_to_csv(self, sheet_models, filepath):
        """Fallback CSV export"""
        try:
            import csv
            
            with open(filepath, 'wb') as f:
                writer = csv.writer(f)
                
                # Header
                writer.writerow([
                    "Sheet Number",
                    "Sheet Name",
                    "Designed By", 
                    "Checked By",
                    "Drawn By",
                    "Approved By"
                ])
                
                # Data
                for sheet_model in sheet_models:
                    writer.writerow([
                        sheet_model.sheet_number.encode('utf-8') if isinstance(sheet_model.sheet_number, unicode) else sheet_model.sheet_number,
                        sheet_model.sheet_name.encode('utf-8') if isinstance(sheet_model.sheet_name, unicode) else sheet_model.sheet_name,
                        sheet_model.designed_by.encode('utf-8') if isinstance(sheet_model.designed_by, unicode) else sheet_model.designed_by,
                        sheet_model.checked_by.encode('utf-8') if isinstance(sheet_model.checked_by, unicode) else sheet_model.checked_by,
                        sheet_model.drawn_by.encode('utf-8') if isinstance(sheet_model.drawn_by, unicode) else sheet_model.drawn_by,
                        sheet_model.approved_by.encode('utf-8') if isinstance(sheet_model.approved_by, unicode) else sheet_model.approved_by
                    ])
            
            return True
        except Exception as e:
            print("Error exporting to CSV: {}".format(str(e)))
            return False
    
    def import_sheets_from_excel(self, filepath):
        """Import sheet data from Excel"""
        try:
            import openpyxl
            
            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            
            # Read data (skip header row)
            sheet_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Has sheet number
                    sheet_data.append({
                        'sheet_number': row[0],
                        'sheet_name': row[1] if row[1] else "",
                        'designed_by': row[2] if len(row) > 2 and row[2] else "-",
                        'checked_by': row[3] if len(row) > 3 and row[3] else "-",
                        'drawn_by': row[4] if len(row) > 4 and row[4] else "-",
                        'approved_by': row[5] if len(row) > 5 and row[5] else "-"
                    })
            
            return sheet_data

        except Exception as e:
            print("Error importing from Excel: {}".format(str(e)))
            import traceback
            traceback.print_exc()
            return None

    # ── Dict-based API used by the Sheet Manager dialog (ManaSheetsDialog) ──────
    # The dialog passes/consumes lists of dicts (not SheetModel objects) and
    # matches rows back to sheets by ElementId, so it needs these two adapters
    # in addition to the model-based methods above. Kept separate so the other
    # caller (import_excel_dialog.py -> import_sheets_from_excel) is untouched.

    def export_sheets(self, filepath, sheet_dicts):
        """Export a list of sheet dicts to Excel (ManaSheets dialog format).

        `sheet_dicts` items carry the keys in ``_EXPORT_COLUMNS`` (including a
        numeric ``id``). Returns True on success, False otherwise; falls back
        to CSV when openpyxl is unavailable.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet List"

            header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
            header_font = Font(bold=True, color="000000")

            for col, (label, _key) in enumerate(self._EXPORT_COLUMNS, 1):
                cell = ws.cell(row=1, column=col, value=label)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row, data in enumerate(sheet_dicts, 2):
                for col, (_label, key) in enumerate(self._EXPORT_COLUMNS, 1):
                    ws.cell(row=row, column=col, value=data.get(key))

            for col in range(1, len(self._EXPORT_COLUMNS) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

            wb.save(filepath)
            return True

        except ImportError:
            return self._export_dicts_to_csv(sheet_dicts, filepath.replace('.xlsx', '.csv'))
        except Exception as e:
            print("Error exporting sheets to Excel: {}".format(str(e)))
            import traceback
            traceback.print_exc()
            return False

    def _export_dicts_to_csv(self, sheet_dicts, filepath):
        """CSV fallback for export_sheets when openpyxl is unavailable."""
        try:
            import csv
            with open(filepath, 'wb') as f:
                writer = csv.writer(f)
                writer.writerow([label for label, _key in self._EXPORT_COLUMNS])
                for data in sheet_dicts:
                    row_vals = []
                    for _label, key in self._EXPORT_COLUMNS:
                        v = data.get(key, "")
                        if isinstance(v, unicode):
                            v = v.encode('utf-8')
                        row_vals.append(v)
                    writer.writerow(row_vals)
            return True
        except Exception as e:
            print("Error exporting sheets to CSV: {}".format(str(e)))
            return False

    def import_sheets(self, filepath):
        """Import sheet rows for the ManaSheets dialog.

        Resolves columns from the header row (so the trailing ID column and any
        column reordering are tolerated) and returns a list of dicts keyed by
        the dialog's field names, with an integer ``id`` when present. Returns
        None on failure, [] when the file has no recognisable sheet columns.
        """
        try:
            import openpyxl

            wb = openpyxl.load_workbook(filepath)
            ws = wb.active
            rows = ws.iter_rows(values_only=True)

            try:
                header = next(rows)
            except StopIteration:
                return []

            col_of = {}
            for idx, label in enumerate(header):
                if label is None:
                    continue
                key = self._HEADER_TO_KEY.get(unicode(label).strip().lower())
                if key is not None:
                    col_of[key] = idx
            if "id" not in col_of and "sheet_number" not in col_of:
                return []   # not a recognisable sheet export

            def _cell(row, key):
                i = col_of.get(key)
                if i is None or i >= len(row):
                    return None
                return row[i]

            result = []
            for row in rows:
                if not row or all(c is None for c in row):
                    continue
                entry = {}
                for key in ("sheet_number", "sheet_name", "designed_by",
                            "checked_by", "drawn_by", "approved_by"):
                    v = _cell(row, key)
                    if v is not None:
                        entry[key] = v
                raw_id = _cell(row, "id")
                if raw_id is not None:
                    try:
                        entry["id"] = int(raw_id)
                    except Exception:
                        pass
                if entry.get("id") is not None or entry.get("sheet_number"):
                    result.append(entry)
            return result

        except Exception as e:
            print("Error importing sheets from Excel: {}".format(str(e)))
            import traceback
            traceback.print_exc()
            return None
